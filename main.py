from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler
import logging
import config  # Файл с токеном бота
import openpyxl
from openpyxl.styles import PatternFill
import os
import shutil  # Для создания резервной копии
import asyncio

# Логирование
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)

# Указываем путь к файлу Excel в корневой папке проекта
excel_file_path = './guest_list.xlsx'

# Список разрешённых пользователей (User ID)
authorized_usernames = ["@Alexander_Dym", "@yana_nebo", "@kasper94", "@AliNadya", "@kastorkin"]

# Список участников с уникальными токенами
participants = {
    "token_kostiantyn": "Kostiantyn Dotsenko",
    "token_mariia": "Mariia",
    "token_joanna": "Joanna Marinova",
    "token_daiana_b": "Daiana",
    "token_daiana_imlive": "Daiana Imlive",
    "token_artem": "Artem",
    "token_anastasiia": "Anastasiia",
    "token_olga": "Olga",
    "token_sebastien": "Sebastien BALESTAS",
    "token_nikita": "Nikita",
    "token_anton": "Anton",
    "token_sasha": "Sasha",
    "token_shawn": "Shawn Herron",
    "token_cybermike": "cybermike",
    "token_vitaliy": "Vitaliy",
    "token_nikolay": "Николай",
    "token_vladyslav": "Vladyslav Haiduk",
    "token_slava": "Slava",
    "token_roman": "Roman",
    "token_tomas": "Tomáš Vítek",
    "token_pavel": "Pavel",
    "token_kira": "Kira",
    "token_snezhana": "Snezhana Fed",
    "token_maksym": "Maksym",
    "token_igor": "Igor Aleksandrov",
    "token_andrii": "Andrii Cherepanskyi",
    "token_nazar": "Nazar",
    "token_sebastian": "Sebastian Prekop",
    "token_tobias_andersen": "Tobias Andersen",
    "token_daria": "Daria",
    "token_kian": "Kian Mir",
    "token_alessandro": "Alessandro Polidoro",
    "token_tobias_endorphina": "Tobias Andersen",
    "token_andrii_stat": "Andrii",
    "token_inna": "Inna",
    "token_muhammad": "Muhammad Ulil Nuha",
    "token_pentil": "Pentil_dawa",
    "token_sergey": "Sergey",
    "token_ilya": "Ilya",
    "token_alexander": "Alexander",
    "token_arthur": "Arthur"
}

# Хранилище для отслеживания состояния участников
checked_in = set()

# Функция для проверки прав доступа по username
def is_authorized(username):
    return username in authorized_usernames

# Функция для отметки строки участника в таблице Excel зеленым цветом
def mark_guest_in_excel(participant_name):
    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb.active

    fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
        if row[0].value == participant_name:
            for cell in row:
                cell.fill = fill
            break

    wb.save(excel_file_path)

# Функция для сброса зелёной заливки в таблице Excel
def reset_excel_fill():
    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb.active

    # Проходим по всем строкам и убираем заливку
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            cell.fill = PatternFill(fill_type=None)  # Убираем заливку

    wb.save(excel_file_path)

# Функция для создания резервной копии Excel файла
def create_backup():
    backup_file_path = excel_file_path.replace(".xlsx", "_backup.xlsx")
    shutil.copyfile(excel_file_path, backup_file_path)
    return backup_file_path

# Функция обработки команды /start
async def start(update: Update, context):
    user = update.message.from_user
    username = f"@{user.username}"

    if not is_authorized(username):
        await update.message.reply_text("❌ У вас нет доступа к этому боту.")
        return

    if context.args:
        parameter = context.args[0]
        logging.info(f"Received parameter: {parameter}")

        if parameter in participants:
            participant_name = participants[parameter]
            if participant_name not in checked_in:
                checked_in.add(participant_name)
                mark_guest_in_excel(participant_name)
                await update.message.reply_text(
                    f"✅ 🎉 *{participant_name}* пришел на мероприятие! 🥳",
                    parse_mode='Markdown'
                )
            else:
                await update.message.reply_text(
                    f"🚫 *{participant_name}* уже зарегистрирован. 🚫",
                    parse_mode='Markdown'
                )
        else:
            await update.message.reply_text("❌ Этот билет не найден. ❌")
    else:
        await menu(update, context)

# Функция для подтверждения сброса данных
async def reset(update: Update, context):
    user = update.message.from_user
    username = f"@{user.username}"

    if not is_authorized(username):
        await update.message.reply_text("❌ У вас нет доступа к этому боту.")
        return

    # Отправляем запрос на подтверждение сброса данных
    keyboard = [[InlineKeyboardButton("ОК ✅", callback_data='confirm_reset')]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text(
        "Вы действительно хотите обнулить данные? Все данные будут сброшены, а резервная копия будет создана.",
        reply_markup=reply_markup
    )

## Функция обработки подтверждения сброса
async def confirm_reset(query, context):
    user = query.from_user
    username = f"@{user.username}"

    if not is_authorized(username):
        await query.message.reply_text("❌ У вас нет доступа к этому боту.")
        return

    # Создание резервной копии файла
    backup_file_path = create_backup()
    
    # Отправка резервной копии пользователю
    await query.message.reply_document(open(backup_file_path, 'rb'), caption="Резервная копия")

    # Сброс данных
    checked_in.clear()
    reset_excel_fill()  # Сбрасываем зелёную заливку в таблице
    
    # Удаляем сообщение с запросом
    await query.message.delete()

    # Уведомление о сбросе
    await query.message.reply_text("🔄 Данные успешно сброшены.")
    
# Функция отправки меню с кнопками
async def menu(update: Update, context):
    user = update.message.from_user
    username = f"@{user.username}"

    if not is_authorized(username):
        await update.message.reply_text("❌ У вас нет доступа к этому боту.")
        return

    keyboard = [
        [InlineKeyboardButton("Guest List", callback_data='guest_list')]  # Оставляем только кнопку "Guest List"
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text("Выберите действие:", reply_markup=reply_markup)

# Функция обработки нажатий на кнопки
async def button(update: Update, context):
    query = update.callback_query
    data = query.data

    if data == 'confirm_reset':
        await confirm_reset(query, context)  # Обрабатываем нажатие на "ОК"
    elif data == 'guest_list':
        await send_excel(query, context)  # Отправляем файл Excel при нажатии на "Guest List"

# Функция для отправки файла Excel
async def send_excel(update: Update, context):
    user = update.message.from_user
    username = f"@{user.username}"

    if not is_authorized(username):
        await update.message.reply_text("❌ У вас нет доступа к этому боту.")
        return

    try:
        await update.message.reply_document(open(excel_file_path, 'rb'))
    except Exception as e:
        await update.message.reply_text(f"Ошибка отправки файла: {str(e)}")

# Обработчик команды /guest_list
async def guest_list(update: Update, context):
    user = update.message.from_user
    username = f"@{user.username}"

    if not is_authorized(username):
        await update.message.reply_text("❌ У вас нет доступа к этому боту.")
        return

    await send_excel(update, context)

def main():
    application = Application.builder().token(config.tg_bot_token).build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("reset", reset))  # Команда для сброса данных
    application.add_handler(CommandHandler("menu", menu))
    application.add_handler(CommandHandler("guest_list", guest_list))  # Команда для отправки файла Excel через команду
    application.add_handler(CallbackQueryHandler(button))  # Обработчик кнопок

    application.run_polling()

if __name__ == '__main__':
    main()